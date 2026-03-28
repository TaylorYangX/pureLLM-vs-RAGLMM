"""
===========================================================
数据加载模块 (Dataset Loader Module)
===========================================================

功能说明：
    负责加载和预处理实验所需的所有数据，包括：
    1. PDF 文档加载 —— 从 PDF 文件中提取纯文本
    2. Excel 文档加载 —— 从 .xlsx 文件中提取文本内容
    3. 多文件批量加载 —— 自动扫描 data/ 目录中所有支持的文件
    4. 文档切分 —— 将长文本切分为适合向量检索的小块
    5. Ground Truth 加载 —— 加载查询和标准答案数据

设计说明：
    本项目支持任意领域的知识文档，不限于特定教科书。
    用户只需将 PDF 或 XLSX 文件放入 data/ 目录即可。
    文档被切分为小块后进行向量化，存入 FAISS 索引（VectorDB/）。
    切分时使用重叠（overlap）确保跨块边界的信息不丢失。

支持的文件格式：
    - .pdf  —— 使用 PyPDF2 提取文本
    - .xlsx —— 使用 openpyxl 提取单元格文本
"""

import json
import os
import glob
from typing import Optional

import PyPDF2


# =============================================
# 1. PDF 文件加载
# =============================================

def load_pdf_document(pdf_path: str) -> str:
    """
    从 PDF 文件中提取全部文本内容。

    原理：
        使用 PyPDF2 逐页读取 PDF 内容。PyPDF2 是纯 Python 实现，
        无需外部系统依赖（如 poppler），适合跨平台使用。
        提取的文本将用于后续的文档切分和向量化。

    参数:
        pdf_path (str): PDF 文件的路径

    返回:
        str: 提取的全部文本内容（各页文本用换行符连接）

    异常:
        FileNotFoundError: 当指定的 PDF 文件不存在时
        RuntimeError: 当 PDF 解析失败时
    """
    # 验证文件是否存在
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"❌ PDF 文件未找到: {pdf_path}")

    print(f"📖 正在加载 PDF: {pdf_path}")

    try:
        # 打开 PDF 文件并逐页提取文本
        with open(pdf_path, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            total_pages = len(reader.pages)
            print(f"   共 {total_pages} 页")

            # 存储每页提取的文本
            all_text = []
            for page_num, page in enumerate(reader.pages):
                # 提取单页文本
                page_text = page.extract_text()
                if page_text:
                    # 清理文本：去除多余空白，但保留段落结构
                    cleaned_text = page_text.strip()
                    all_text.append(cleaned_text)

            # 用换行符连接所有页面的文本
            full_text = "\n\n".join(all_text)
            print(f"   提取文本长度: {len(full_text)} 字符")

            return full_text

    except Exception as e:
        raise RuntimeError(f"❌ PDF 解析失败: {e}")


# =============================================
# 2. Excel 文件加载
# =============================================

def load_xlsx_document(xlsx_path: str) -> str:
    """
    从 Excel (.xlsx) 文件中提取全部文本内容。

    原理：
        使用 openpyxl 读取 .xlsx 文件中的每个工作表、每个单元格。
        将所有非空单元格的文本内容按行拼接，工作表之间用分隔符区分。
        适用于结构化数据（如表格形式的知识库）和半结构化文本。

    处理策略：
        - 遍历所有工作表（sheet），不遗漏数据
        - 每行的单元格用制表符连接，保留表格结构信息
        - 空行和空单元格自动跳过
        - 工作表名称作为标题，便于后续检索时定位来源

    参数:
        xlsx_path (str): Excel 文件的路径

    返回:
        str: 提取的全部文本内容

    异常:
        FileNotFoundError: 当指定的文件不存在时
        RuntimeError: 当解析失败时
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"❌ Excel 文件未找到: {xlsx_path}")

    # openpyxl 不在顶层导入，因为它是可选依赖
    # 只有在实际需要处理 xlsx 时才加载
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "❌ 处理 .xlsx 文件需要 openpyxl 库。\n"
            "请运行: pip install openpyxl"
        )

    print(f"📊 正在加载 Excel: {xlsx_path}")

    try:
        # 以只读模式打开，减少内存占用
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        print(f"   共 {len(sheet_names)} 个工作表: {sheet_names}")

        all_text_parts = []

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            # 添加工作表标题作为上下文标记
            sheet_lines = [f"=== 工作表: {sheet_name} ==="]

            for row in sheet.iter_rows(values_only=True):
                # 将每行中的非空单元格转换为字符串并连接
                row_values = []
                for cell in row:
                    if cell is not None:
                        row_values.append(str(cell).strip())

                # 跳过完全空的行
                if row_values:
                    sheet_lines.append("\t".join(row_values))

            # 只有在工作表有内容时才添加
            if len(sheet_lines) > 1:  # 大于1是因为第一行是标题
                all_text_parts.append("\n".join(sheet_lines))

        workbook.close()

        full_text = "\n\n".join(all_text_parts)
        print(f"   提取文本长度: {len(full_text)} 字符")

        return full_text

    except Exception as e:
        raise RuntimeError(f"❌ Excel 解析失败: {e}")


# =============================================
# 3. 通用文档加载器
# =============================================

def load_document(file_path: str) -> str:
    """
    根据文件扩展名自动选择合适的加载器。

    支持的格式：
        - .pdf  → load_pdf_document()
        - .xlsx → load_xlsx_document()

    参数:
        file_path (str): 文件路径

    返回:
        str: 提取的文本内容
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        return load_pdf_document(file_path)
    elif ext == ".xlsx":
        return load_xlsx_document(file_path)
    else:
        raise ValueError(
            f"❌ 不支持的文件格式: {ext}\n"
            f"支持的格式: .pdf, .xlsx"
        )


def load_all_documents(data_dir: str = "data") -> str:
    """
    扫描目录，加载所有支持格式的文档并合并为一个文本。

    原理：
        用户可以将多个文档放入 data/ 目录，本函数会自动发现
        并加载所有 PDF 和 XLSX 文件。多个文档的文本合并后
        统一进行切分和向量化。

    参数:
        data_dir (str): 文档目录路径，默认 "data"

    返回:
        str: 所有文档合并后的文本

    异常:
        FileNotFoundError: 当目录中没有找到任何支持的文件时
    """
    # 搜索所有支持的文件
    supported_extensions = ["*.pdf", "*.xlsx"]
    all_files = []
    for ext_pattern in supported_extensions:
        found = glob.glob(os.path.join(data_dir, ext_pattern))
        all_files.extend(found)

    # 也搜索子目录（一级）
    for ext_pattern in supported_extensions:
        found = glob.glob(os.path.join(data_dir, "*", ext_pattern))
        all_files.extend(found)

    # 去重并排序
    all_files = sorted(set(all_files))

    if not all_files:
        raise FileNotFoundError(
            f"❌ 在 {data_dir}/ 目录中未找到任何支持的文件。\n"
            f"请将 PDF 或 XLSX 文件放入 {data_dir}/ 目录。\n"
            f"支持的格式: .pdf, .xlsx"
        )

    print(f"📂 在 {data_dir}/ 中找到 {len(all_files)} 个文档:")
    for f in all_files:
        size_mb = os.path.getsize(f) / 1024 / 1024
        print(f"   - {os.path.basename(f)} ({size_mb:.1f} MB)")

    # 逐个加载并合并
    all_texts = []
    for file_path in all_files:
        try:
            text = load_document(file_path)
            # 添加文件来源标记，方便后续追溯
            header = f"\n{'='*60}\n📄 文件来源: {os.path.basename(file_path)}\n{'='*60}\n"
            all_texts.append(header + text)
        except Exception as e:
            print(f"⚠️  加载失败，跳过: {file_path} ({e})")

    if not all_texts:
        raise RuntimeError("❌ 所有文档加载失败，无可用文本。")

    merged_text = "\n\n".join(all_texts)
    print(f"\n📝 总文本长度: {len(merged_text)} 字符")

    return merged_text


# =============================================
# 4. 文档切分
# =============================================

def split_documents(
    text: str,
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
    separator: str = "\n"
) -> list:
    """
    将长文本切分为适合向量检索的小块。

    原理：
        RAG 的核心步骤之一。长文本无法直接输入 embedding 模型，
        需要切分为较小的"块"（chunk）。切分策略直接影响检索质量：

        1. chunk_size 控制每块的大小：
           - 太大：包含过多无关信息，降低检索精度
           - 太小：丢失上下文，导致信息碎片化

        2. chunk_overlap 控制相邻块的重叠：
           - 重叠确保跨块边界的完整句子不被截断
           - 提高检索时找到完整相关段落的概率

        3. separator 优先在自然边界（如换行符）处切分：
           - 避免在句子中间断开
           - 保持语义完整性

    参数:
        text (str): 待切分的完整文本
        chunk_size (int): 每个块的目标字符数，默认 1000
        chunk_overlap (int): 相邻块间重叠的字符数，默认 200
        separator (str): 优先切分的分隔符，默认 "\n"

    返回:
        list: 切分后的文本块列表，每个元素为一个字符串
    """
    # 参数验证
    if chunk_overlap >= chunk_size:
        raise ValueError(
            f"chunk_overlap ({chunk_overlap}) 必须小于 chunk_size ({chunk_size})"
        )

    # 步骤1: 按分隔符初步分割文本
    # 这样可以优先在自然段落边界处切分
    splits = text.split(separator)

    chunks = []          # 最终的文本块列表
    current_chunk = []   # 当前正在构建的块（存储分割后的片段）
    current_length = 0   # 当前块的累计字符数

    for split in splits:
        split_length = len(split)

        # 判断是否需要开始新的块：
        # 当加入当前片段后超过 chunk_size 时，保存当前块并开始新块
        if current_length + split_length + 1 > chunk_size and current_chunk:
            # 将当前块的所有片段合并为一个字符串
            chunk_text = separator.join(current_chunk).strip()
            if chunk_text:
                chunks.append(chunk_text)

            # 实现重叠：保留当前块末尾的部分片段
            # 从后往前累计，直到达到 overlap 的字符数
            overlap_chunks = []
            overlap_length = 0
            for prev_split in reversed(current_chunk):
                if overlap_length + len(prev_split) + 1 <= chunk_overlap:
                    overlap_chunks.insert(0, prev_split)
                    overlap_length += len(prev_split) + 1
                else:
                    break

            # 新块从重叠部分开始
            current_chunk = overlap_chunks
            current_length = overlap_length

        # 将当前片段加入块
        current_chunk.append(split)
        current_length += split_length + 1  # +1 是分隔符的长度

    # 处理最后一个块
    if current_chunk:
        chunk_text = separator.join(current_chunk).strip()
        if chunk_text:
            chunks.append(chunk_text)

    print(f"📦 文档切分完成: {len(chunks)} 个文本块")
    if chunks:
        print(f"   平均块大小: {sum(len(c) for c in chunks) / len(chunks):.0f} 字符")

    return chunks


# =============================================
# 5. Ground Truth 加载
# =============================================

def load_ground_truth(json_path: str) -> list:
    """
    加载 Ground Truth 数据（查询 + 标准答案）。

    数据格式为包含三元组的列表：
        - query_id: 查询编号
        - query: 查询问题
        - ground_truth: 标准答案
        - source_passage: 答案来源的原文段落

    参数:
        json_path (str): Ground Truth JSON 文件的路径

    返回:
        list: 包含查询和答案的字典列表
    """
    if not os.path.exists(json_path):
        raise FileNotFoundError(
            f"❌ Ground Truth 文件未找到: {json_path}\n"
            f"请先运行 step2_generate_ground_truth.py 生成，\n"
            f"或手动创建 ground_truth.json 文件。"
        )

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    print(f"📋 加载 Ground Truth: {len(data)} 条查询")
    for item in data:
        print(f"   Q{item['query_id']}: {item['query'][:50]}...")

    return data


# =============================================
# 单元测试 / 功能演示
# =============================================
if __name__ == "__main__":
    # 测试 Ground Truth 加载
    print("测试 Ground Truth 加载...")
    gt_path = os.path.join(os.path.dirname(__file__), "ground_truth.json")
    if os.path.exists(gt_path):
        data = load_ground_truth(gt_path)
        print(f"\n示例查询: {data[0]['query']}")
        print(f"示例答案: {data[0]['ground_truth'][:100]}...")
    else:
        print(f"⚠️  {gt_path} 不存在，跳过测试")

    # 测试文本切分
    print("\n\n测试文本切分...")
    sample_text = "This is a test.\n" * 100
    chunks = split_documents(sample_text, chunk_size=200, chunk_overlap=50)
    print(f"输入长度: {len(sample_text)}, 输出块数: {len(chunks)}")

    # 测试文档发现
    print("\n\n测试文档发现...")
    try:
        text = load_all_documents("data")
        print(f"合并文本长度: {len(text)}")
    except FileNotFoundError as e:
        print(f"⚠️  {e}")
